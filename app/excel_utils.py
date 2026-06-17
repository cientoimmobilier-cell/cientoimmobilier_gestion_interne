from io import BytesIO
import openpyxl
from openpyxl import Workbook
from app.models import Client, Propriete, Proprietaire

def export_clients_to_excel(clients):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    headers = [
        "Code Client", "Nom", "Prénom", "Téléphone", 
        "Téléphone 2", "E-mail", "Adresse", "Ville", 
        "Profession", "Budget Min", "Budget Max", "Source", "Observations"
    ]
    ws.append(headers)
    
    for client in clients:
        ws.append([
            client.code_client or "",
            client.nom or "",
            client.prenom or "",
            client.telephone or "",
            client.telephone_secondaire or "",
            client.email or "",
            client.adresse or "",
            client.ville or "",
            client.profession or "",
            float(client.budget_min) if client.budget_min is not None else "",
            float(client.budget_max) if client.budget_max is not None else "",
            client.source_client or "",
            client.observations or ""
        ])
        
    # Appliquer un style simple (en-tête en gras)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def import_clients_from_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active
    
    clients_data = []
    # Parcourir en ignorant la première ligne (les en-têtes)
    iter_rows = ws.iter_rows(min_row=2, values_only=True)
    for row in iter_rows:
        if not row or not any(row):
            continue
            
        # Mapper les données
        client_dict = {
            "nom": str(row[1]).strip().upper() if row[1] else "",
            "prenom": str(row[2]).strip().title() if row[2] else "",
            "telephone": str(row[3]).strip() if row[3] else "",
            "telephone_secondaire": str(row[4]).strip() if row[4] else "",
            "email": str(row[5]).strip() if row[5] else "",
            "adresse": str(row[6]).strip() if row[6] else "",
            "ville": str(row[7]).strip() if row[7] else "",
            "profession": str(row[8]).strip() if row[8] else "",
            "budget_min": float(row[9]) if row[9] is not None and str(row[9]).strip() != "" else None,
            "budget_max": float(row[10]) if row[10] is not None and str(row[10]).strip() != "" else None,
            "source_client": str(row[11]).strip() if row[11] else "Import Excel",
            "observations": str(row[12]).strip() if row[12] else ""
        }
        
        # Ignorer si le nom est vide
        if client_dict["nom"]:
            clients_data.append(client_dict)
            
    return clients_data

def export_properties_to_excel(properties):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proprietes"
    
    headers = [
        "Référence", "Titre", "Type de bien", "Opération", 
        "Adresse", "Ville", "Quartier", "Prix", "Devise",
        "Superficie (m²)", "Chambres", "Salles de bain", "Garages", 
        "Statut", "Propriétaire E-mail", "Description"
    ]
    ws.append(headers)
    
    for prop in properties:
        owner_email = prop.proprietaire.email if prop.proprietaire else ""
        ws.append([
            prop.reference_bien or "",
            prop.titre or "",
            prop.type_bien or "",
            prop.type_operation or "",
            prop.adresse or "",
            prop.ville or "",
            prop.quartier or "",
            float(prop.prix) if prop.prix is not None else 0.0,
            prop.devise or "EUR",
            float(prop.superficie) if prop.superficie is not None else "",
            prop.nombre_chambres if prop.nombre_chambres is not None else "",
            prop.nombre_salles_bain if prop.nombre_salles_bain is not None else "",
            prop.nombre_garages if prop.nombre_garages is not None else "",
            prop.statut or "Disponible",
            owner_email,
            prop.description or ""
        ])
        
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def import_properties_from_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active
    
    properties_data = []
    iter_rows = ws.iter_rows(min_row=2, values_only=True)
    for row in iter_rows:
        if not row or not any(row):
            continue
            
        # Rechercher le propriétaire par email si renseigné
        owner_id = None
        owner_email = str(row[14]).strip() if len(row) > 14 and row[14] else None
        if owner_email:
            owner = Proprietaire.query.filter_by(email=owner_email).first()
            if owner:
                owner_id = owner.id
                
        prop_dict = {
            "titre": str(row[1]).strip() if row[1] else "",
            "type_bien": str(row[2]).strip() if row[2] else "Appartement",
            "type_operation": str(row[3]).strip() if row[3] else "Vente",
            "adresse": str(row[4]).strip() if row[4] else "",
            "ville": str(row[5]).strip() if row[5] else "",
            "quartier": str(row[6]).strip() if row[6] else "",
            "prix": float(row[7]) if row[7] is not None else 0.0,
            "devise": str(row[8]).strip() if len(row) > 8 and row[8] else "EUR",
            "superficie": float(row[9]) if len(row) > 9 and row[9] is not None and str(row[9]).strip() != "" else None,
            "nombre_chambres": int(row[10]) if len(row) > 10 and row[10] is not None and str(row[10]).strip() != "" else None,
            "nombre_salles_bain": int(row[11]) if len(row) > 11 and row[11] is not None and str(row[11]).strip() != "" else None,
            "nombre_garages": int(row[12]) if len(row) > 12 and row[12] is not None and str(row[12]).strip() != "" else None,
            "statut": str(row[13]).strip() if len(row) > 13 and row[13] else "Disponible",
            "proprietaire_id": owner_id,
            "description": str(row[15]).strip() if len(row) > 15 and row[15] else ""
        }
        
        if prop_dict["titre"] and prop_dict["prix"] > 0:
            properties_data.append(prop_dict)
            
    return properties_data
