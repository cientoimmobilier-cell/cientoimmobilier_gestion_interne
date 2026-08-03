from io import BytesIO

import openpyxl
from openpyxl import Workbook
from sqlalchemy import select

from app import db
from app.models import Proprietaire
from app.utils.helpers import neutralize_formula

def _safe(value):
    return neutralize_formula(value)


def _to_number(value):
    """Convertit une cellule Excel en nombre, sans jamais lever d'exception.

    Accepte les formats courants : ``1500``, ``1 500``, ``1 500,50 €``, ``1,5``.
    Retourne None si la valeur n'est pas convertible (ligne ignorée au lieu
    de faire planter tout l'import).
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return None
    cleaned = raw.replace('\u00a0', ' ').replace(' ', '').replace('€', '')
    cleaned = cleaned.replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        return None

def export_clients_to_excel(clients):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients'

    headers = [
        'Code Client', 'Nom', 'Prénom', 'Téléphone',
        'Téléphone 2', 'E-mail', 'Adresse', 'Ville',
        'Profession', 'Zone Ciblée', 'Description',
        'Budget Min', 'Budget Max', 'Source', 'Observations'
    ]
    ws.append(headers)
    
    for client in clients:
        ws.append([
            _safe(client.code_client or ''),
            _safe(client.nom or ''),
            _safe(client.prenom or ''),
            _safe(client.telephone or ''),
            _safe(client.telephone_secondaire or ''),
            _safe(client.email or ''),
            _safe(client.adresse or ''),
            _safe(client.ville or ''),
            _safe(client.profession or ''),
            _safe(client.zone_ciblee or ''),
            _safe(client.description or ''),
            float(client.budget_min) if client.budget_min is not None else '',
            float(client.budget_max) if client.budget_max is not None else '',
            _safe(client.source_client or ''),
            _safe(client.observations or '')
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_clients_from_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active

    clients_data = []
    iter_rows = ws.iter_rows(min_row=2, values_only=True)
    for row in iter_rows:
        if not row or not any(row):
            continue

        row_len = len(row)

        client_dict = {
            'nom': str(row[1]).strip().upper() if row_len > 1 and row[1] else '',
            'prenom': str(row[2]).strip().title() if row_len > 2 and row[2] else '',
            'telephone': str(row[3]).strip() if row_len > 3 and row[3] else '',
            'telephone_secondaire': str(row[4]).strip() if row_len > 4 and row[4] else '',
            'email': str(row[5]).strip() if row_len > 5 and row[5] else '',
            'adresse': str(row[6]).strip() if row_len > 6 and row[6] else '',
            'ville': str(row[7]).strip() if row_len > 7 and row[7] else '',
            'profession': str(row[8]).strip() if row_len > 8 and row[8] else '',
            'zone_ciblee': str(row[9]).strip() if row_len > 9 and row[9] else '',
            'description': str(row[10]).strip() if row_len > 10 and row[10] else '',
            'budget_min': _to_number(row[11]) if row_len > 11 else None,
            'budget_max': _to_number(row[12]) if row_len > 12 else None,
            'source_client': str(row[13]).strip() if row_len > 13 and row[13] else 'Import Excel',
            'observations': str(row[14]).strip() if row_len > 14 and row[14] else '',
        }

        if client_dict['nom']:
            clients_data.append(client_dict)

    return clients_data


def export_properties_to_excel(properties):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proprietes'

    headers = [
        'Référence', 'Titre', 'Type de bien', 'Opération',
        'Adresse', 'Ville', 'Quartier', 'Prix', 'Devise',
        'Superficie (m²)', 'Chambres', 'Salles de bain', 'Garages',
        'Statut', 'Propriétaire E-mail', 'Description'
    ]
    ws.append(headers)
    
    for prop in properties:
        owner_email = prop.proprietaire.email if prop.proprietaire else ''
        ws.append([
            _safe(prop.reference_bien or ''),
            _safe(prop.titre or ''),
            _safe(prop.type_bien or ''),
            _safe(prop.type_operation or ''),
            _safe(prop.adresse or ''),
            _safe(prop.ville or ''),
            _safe(prop.quartier or ''),
            float(prop.prix) if prop.prix is not None else 0.0,
            _safe(prop.devise or 'EUR'),
            float(prop.superficie) if prop.superficie is not None else '',
            prop.nombre_chambres if prop.nombre_chambres is not None else '',
            prop.nombre_salles_bain if prop.nombre_salles_bain is not None else '',
            prop.nombre_garages if prop.nombre_garages is not None else '',
            _safe(prop.statut or 'Disponible'),
            _safe(owner_email),
            _safe(prop.description or '')
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_properties_from_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active

    properties_data = []
    iter_rows = ws.iter_rows(min_row=2, values_only=True)
    for row in iter_rows:
        if not row or not any(row):
            continue

        owner_id = None
        owner_email = str(row[14]).strip() if len(row) > 14 and row[14] else None
        if owner_email:
            owner = db.session.execute(select(Proprietaire).where(Proprietaire.email == owner_email)).scalars().first()
            if owner:
                owner_id = owner.id

        prop_dict = {
            'titre': str(row[1]).strip() if row[1] else '',
            'type_bien': str(row[2]).strip() if row[2] else 'Appartement',
            'type_operation': str(row[3]).strip() if row[3] else 'Vente',
            'adresse': str(row[4]).strip() if row[4] else '',
            'ville': str(row[5]).strip() if row[5] else '',
            'quartier': str(row[6]).strip() if row[6] else '',
            'prix': _to_number(row[7]) if len(row) > 7 else None,
            'devise': str(row[8]).strip() if len(row) > 8 and row[8] else 'EUR',
            'superficie': _to_number(row[9]) if len(row) > 9 else None,
            'nombre_chambres': _to_number(row[10]) if len(row) > 10 else None,
            'nombre_salles_bain': _to_number(row[11]) if len(row) > 11 else None,
            'nombre_garages': _to_number(row[12]) if len(row) > 12 else None,
            'statut': str(row[13]).strip() if len(row) > 13 and row[13] else 'Disponible',
            'proprietaire_id': owner_id,
            'description': str(row[15]).strip() if len(row) > 15 and row[15] else '',
        }

        if prop_dict['titre'] and prop_dict['prix'] is not None and prop_dict['prix'] > 0:
            properties_data.append(prop_dict)

    return properties_data


def export_owners_to_excel(owners):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proprietaires'

    headers = [
        'Nom', 'Prénom', 'Téléphone', 'E-mail',
        'Adresse', 'N° Identité', 'Observations'
    ]
    ws.append(headers)

    for owner in owners:
        ws.append([
            _safe(owner.nom or ''),
            _safe(owner.prenom or ''),
            _safe(owner.telephone or ''),
            _safe(owner.email or ''),
            _safe(owner.adresse or ''),
            _safe(owner.numero_identite or ''),
            _safe(owner.observations or '')
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_owners_from_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active

    owners_data = []
    iter_rows = ws.iter_rows(min_row=2, values_only=True)
    for row in iter_rows:
        if not row or not any(row):
            continue

        owner_dict = {
            'nom': str(row[0]).strip().upper() if row[0] else '',
            'prenom': str(row[1]).strip().title() if row[1] else '',
            'telephone': str(row[2]).strip() if len(row) > 2 and row[2] else '',
            'email': str(row[3]).strip() if len(row) > 3 and row[3] else '',
            'adresse': str(row[4]).strip() if len(row) > 4 and row[4] else '',
            'numero_identite': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            'observations': str(row[6]).strip() if len(row) > 6 and row[6] else '',
        }

        if owner_dict['nom']:
            owners_data.append(owner_dict)

    return owners_data


def export_transactions_to_excel(transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    headers = [
        'Référence', 'Type', 'Bien', 'Client',
        'Agent', 'Montant', 'Devise', 'Date',
        'Statut', 'Observations'
    ]
    ws.append(headers)
    
    for tx in transactions:
        ws.append([
            _safe(tx.reference_transaction or ''),
            _safe(tx.type_transaction or ''),
            _safe(tx.propriete.reference_bien) if tx.propriete else '',
            _safe(f"{tx.client.prenom} {tx.client.nom}") if tx.client else '',
            _safe(f"{tx.agent.prenom} {tx.agent.nom}") if tx.agent else '',
            float(tx.montant) if tx.montant is not None else 0.0,
            _safe(tx.devise or 'EUR'),
            tx.date_transaction.strftime('%d/%m/%Y') if tx.date_transaction else '',
            _safe(tx.statut or ''),
            _safe(tx.observations or '')
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_occupations_to_excel(occupations):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Occupations'

    headers = [
        'N\u00b0 Occupation', 'Statut', 'Client', 'Bien', 'Agent',
        'Date entr\u00e9e', 'Date sortie pr\u00e9vue', 'Date sortie r\u00e9elle',
        'Nombre occupants', 'Contrat', 'Observations'
    ]
    ws.append(headers)

    for occ in occupations:
        client_name = f'{occ.client.prenom or ""} {occ.client.nom or ""}'.strip() if occ.client else ''
        bien_ref = occ.propriete.reference_bien if occ.propriete else ''
        agent_name = f'{occ.agent.prenom or ""} {occ.agent.nom or ""}'.strip() if occ.agent else ''
        contrat = occ.contrat.numero_contrat if occ.contrat else ''

        ws.append([
            _safe(occ.numero_occupation or ''),
            _safe(occ.statut or ''),
            _safe(client_name),
            _safe(bien_ref),
            _safe(agent_name),
            occ.date_entree.strftime('%d/%m/%Y') if occ.date_entree else '',
            occ.date_sortie_prevue.strftime('%d/%m/%Y') if occ.date_sortie_prevue else '',
            occ.date_sortie_reelle.strftime('%d/%m/%Y') if occ.date_sortie_reelle else '',
            occ.nombre_occupants,
            _safe(contrat),
            _safe(occ.observations or '')
        ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
